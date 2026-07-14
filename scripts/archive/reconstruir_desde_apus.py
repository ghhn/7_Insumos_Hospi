#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import psycopg2
import sys
sys.stdout.reconfigure(encoding='utf-8')

conn = psycopg2.connect(
    host='localhost',
    database='7_insumos_rado',
    user='postgres',
    password='Jo.9839514500',
    port=5432
)
cur = conn.cursor()

print("="*100)
print("RECONSTRUIR ESTRUCTURA DESDE APUS_EXTRAIDOS_V2.xlsx")
print("="*100)

try:
    cur.execute("BEGIN")

    # PASO 1: Limpiar tablas
    print("\n[1/4] Limpiando tablas...")
    cur.execute("DELETE FROM insumos")
    cur.execute("DELETE FROM partidas")
    print("  ✓ Limpió partidas e insumos")

    # PASO 2: Cargar partidas e insumos desde APUS_Extraidos_v2.xlsx
    print("\n[2/4] Cargando desde APUS_Extraidos_v2.xlsx...")
    wb = openpyxl.load_workbook('APUS_Extraidos_v2.xlsx')
    ws = wb.active

    partidas_dict = {}  # Guardar partidas únicas
    insumos_list = []   # Guardar relaciones partida-insumo

    for row_idx in range(2, ws.max_row + 1):
        try:
            partida_codigo = ws.cell(row_idx, 1).value
            partida_descripcion = ws.cell(row_idx, 2).value
            partida_rendimiento = ws.cell(row_idx, 3).value
            partida_unidad = ws.cell(row_idx, 4).value
            partida_costo = ws.cell(row_idx, 5).value
            tipo_insumo = ws.cell(row_idx, 6).value
            insumo_codigo = ws.cell(row_idx, 7).value
            insumo_descripcion = ws.cell(row_idx, 8).value
            insumo_unidad = ws.cell(row_idx, 9).value
            insumo_recursos = ws.cell(row_idx, 10).value
            insumo_cantidad = ws.cell(row_idx, 11).value
            insumo_precio = ws.cell(row_idx, 12).value
            insumo_parcial = ws.cell(row_idx, 13).value

            if not partida_codigo or not insumo_codigo:
                continue

            # Guardar partida única
            if str(partida_codigo) not in partidas_dict:
                partidas_dict[str(partida_codigo)] = {
                    'codigo': str(partida_codigo).strip(),
                    'descripcion': str(partida_descripcion).strip() if partida_descripcion else '',
                    'rendimiento': str(partida_rendimiento).strip() if partida_rendimiento else '',
                    'unidad': str(partida_unidad).strip() if partida_unidad else '',
                    'costo': float(partida_costo) if partida_costo else 0
                }

            # Guardar relación partida-insumo
            insumos_list.append({
                'partida_codigo': str(partida_codigo).strip(),
                'tipo_insumo': str(tipo_insumo).strip() if tipo_insumo else '',
                'insumo_codigo': str(insumo_codigo).strip() if insumo_codigo else '',
                'insumo_descripcion': str(insumo_descripcion).strip() if insumo_descripcion else '',
                'insumo_unidad': str(insumo_unidad).strip() if insumo_unidad else '',
                'insumo_recursos': float(insumo_recursos) if insumo_recursos else 0,
                'insumo_cantidad': float(insumo_cantidad) if insumo_cantidad else 0,
                'insumo_precio': float(insumo_precio) if insumo_precio else 0,
                'insumo_parcial': float(insumo_parcial) if insumo_parcial else 0
            })

        except Exception as e:
            pass

    print(f"  ✓ Partidas únicas encontradas: {len(partidas_dict)}")
    print(f"  ✓ Relaciones partida-insumo: {len(insumos_list)}")

    # PASO 3: Insertar partidas
    print("\n[3/4] Insertando partidas...")
    partidas_insertadas = 0
    for cod, data in partidas_dict.items():
        cur.execute("""
            INSERT INTO partidas (codigo, descripcion, unidad, metrado_fijo)
            VALUES (%s, %s, %s, %s)
        """, (data['codigo'], data['descripcion'], data['unidad'], data['costo']))
        partidas_insertadas += 1

    print(f"  ✓ Partidas insertadas: {partidas_insertadas}")

    # PASO 4: Insertar insumos con relación a partidas
    print("\n[4/4] Insertando insumos con relaciones...")
    insumos_insertados = 0
    for insumo in insumos_list:
        try:
            cur.execute("""
                INSERT INTO insumos (
                    codigo_partida, descripcion, unidad,
                    incidencia_original, parcial_original,
                    incidencia, cantidad_modificada, cantidad_adquirida,
                    comentario, es_extra
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                insumo['partida_codigo'],
                insumo['insumo_descripcion'],
                insumo['insumo_unidad'],
                insumo['insumo_cantidad'],
                insumo['insumo_parcial'],
                insumo['insumo_cantidad'],
                insumo['insumo_cantidad'],
                0,
                f"{insumo['tipo_insumo']} - Recursos: {insumo['insumo_recursos']}",
                False
            ))
            insumos_insertados += 1
        except:
            pass

    cur.execute("COMMIT")

    print(f"  ✓ Insumos insertados: {insumos_insertados}")

    # VERIFICACIÓN
    print("\n" + "="*100)
    print("VERIFICACIÓN")
    print("="*100)

    cur.execute("SELECT COUNT(*) FROM partidas")
    total_partidas = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM insumos WHERE es_extra = FALSE")
    total_insumos = cur.fetchone()[0]

    print(f"✓ Total partidas: {total_partidas}")
    print(f"✓ Total insumos: {total_insumos}")

    # Mostrar ejemplo
    cur.execute("""
        SELECT DISTINCT p.codigo, p.descripcion, COUNT(i.id) as qty
        FROM partidas p
        LEFT JOIN insumos i ON i.codigo_partida = p.codigo
        GROUP BY p.codigo, p.descripcion
        LIMIT 3
    """)
    print(f"\n✓ Ejemplo de partidas con insumos:")
    for cod, desc, qty in cur.fetchall():
        print(f"  {cod}: {desc[:50]} ({qty} insumos)")

    print("\n" + "="*100)
    print("✓ RECONSTRUCCIÓN COMPLETADA")
    print("="*100)

except Exception as e:
    cur.execute("ROLLBACK")
    print(f"\n❌ ERROR: {e}")
    import traceback
    traceback.print_exc()

cur.close()
conn.close()
