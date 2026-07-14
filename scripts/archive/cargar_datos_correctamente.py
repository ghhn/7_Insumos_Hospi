#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import psycopg2
import sys
sys.stdout.reconfigure(encoding='utf-8')

conn = psycopg2.connect(
    host='localhost',
    database='7_insumos_rado',
    user='postgres',
    password='Jo.9839514500',
    port=5432
)
cur = conn.cursor()

print("="*100)
print("CARGA CORRECTA DE DATOS DESDE NUEVA_DATA.xlsx")
print("="*100)

try:
    cur.execute("BEGIN")

    # PASO 1: Limpiar insumos (pero NO partidas ni compras)
    print("\n[1/5] Limpiando tabla insumos...")
    cur.execute("DELETE FROM insumos")
    print("  ✓ Limpió insumos")

    # PASO 2: Cargar datos de NUEVA_DATA.xlsx
    print("\n[2/5] Cargando NUEVA_DATA.xlsx...")
    wb = openpyxl.load_workbook('NUEVA_DATA.xlsx')
    ws = wb.active

    insumos_cargados = 0
    vinculos_cargados = 0

    for row_idx in range(2, ws.max_row + 1):
        try:
            # Columnas A-H (INSUMO)
            codigo = ws[f'C{row_idx}'].value  # CODIGO (columna C)
            descripcion = ws[f'D{row_idx}'].value  # DESCRIPCION (P)
            unidad = ws[f'E{row_idx}'].value  # UNIDAD (P)
            cantidad = ws[f'F{row_idx}'].value  # CANTIDAD (P)

            if not descripcion or not codigo:
                continue

            # Convertir valores
            codigo_str = str(int(float(codigo))) if codigo else None
            descripcion_str = str(descripcion).strip() if descripcion else None
            unidad_str = str(unidad).strip() if unidad else "UND."
            cantidad_num = float(cantidad) if cantidad else 0

            # Insertar insumo
            cur.execute("""
                INSERT INTO insumos (
                    codigo_partida, descripcion, unidad,
                    incidencia_original, parcial_original,
                    incidencia, cantidad_modificada, cantidad_adquirida,
                    comentario, es_extra
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                codigo_str, descripcion_str, unidad_str,
                cantidad_num, cantidad_num,
                cantidad_num, cantidad_num,
                0, None, False
            ))
            insumos_cargados += 1

            # Columnas I-R (COMPRA) - si tiene vinculación
            orden_doc = ws[f'L{row_idx}'].value  # ORDEN/DOC (columna L = I+3)

            if orden_doc:
                # Buscar compra en la BD por orden_doc
                orden_str = str(int(float(orden_doc))) if orden_doc else None

                cur.execute("SELECT id FROM compras WHERE orden_doc = %s LIMIT 1", (orden_str,))
                compra_result = cur.fetchone()

                if compra_result:
                    compra_id = compra_result[0]
                    # Insertar vínculo
                    cur.execute("""
                        INSERT INTO mapeo_vinculacion (insumo_nombre, compra_id)
                        VALUES (%s, %s)
                        ON CONFLICT (insumo_nombre, compra_id) DO NOTHING
                    """, (descripcion_str, compra_id))
                    if cur.rowcount > 0:
                        vinculos_cargados += 1

        except Exception as e:
            pass  # Continuar con siguiente fila

    print(f"  ✓ Insumos cargados: {insumos_cargados}")
    print(f"  ✓ Vínculos creados: {vinculos_cargados}")

    # PASO 3: Cargar caja chica
    print("\n[3/5] Cargando caja_chica_nuevo.xlsx...")
    try:
        wb_caja = openpyxl.load_workbook('caja_chica_nuevo.xlsx')
        ws_caja = wb_caja.active

        caja_cargados = 0
        for row_idx in range(2, ws_caja.max_row + 1):
            try:
                detalle = ws_caja[f'A{row_idx}'].value  # insumo_descripcion
                unidad = ws_caja[f'B{row_idx}'].value  # unidad_und
                cantidad = ws_caja[f'C{row_idx}'].value  # cantidad_c

                if not detalle:
                    continue

                detalle_str = str(detalle).strip()
                unidad_str = str(unidad).strip() if unidad else "UND."
                cantidad_num = float(cantidad) if cantidad else 0

                cur.execute("""
                    INSERT INTO insumos (
                        codigo_partida, descripcion, unidad,
                        incidencia_original, parcial_original,
                        incidencia, cantidad_modificada, cantidad_adquirida,
                        comentario, es_extra
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    'CJA.CHI', detalle_str, unidad_str,
                    cantidad_num, cantidad_num,
                    cantidad_num, cantidad_num,
                    0, None, False
                ))
                caja_cargados += 1
            except:
                pass

        print(f"  ✓ Caja chica cargados: {caja_cargados}")
    except:
        print("  ⚠️ No se encontró caja_chica_nuevo.xlsx")

    cur.execute("COMMIT")

    # PASO 4: Verificación
    print("\n[4/5] Verificación de datos...")
    cur.execute("SELECT COUNT(*) FROM insumos WHERE es_extra = FALSE")
    total_insumos = cur.fetchone()[0]
    print(f"  ✓ Total insumos: {total_insumos}")

    cur.execute("SELECT COUNT(*) FROM mapeo_vinculacion")
    total_vinculos = cur.fetchone()[0]
    print(f"  ✓ Total vinculaciones: {total_vinculos}")

    cur.execute("SELECT COUNT(*) FROM partidas")
    total_partidas = cur.fetchone()[0]
    print(f"  ✓ Total partidas: {total_partidas}")

    cur.execute("SELECT COUNT(*) FROM compras")
    total_compras = cur.fetchone()[0]
    print(f"  ✓ Total compras: {total_compras}")

    print("\n" + "="*100)
    print("✓ CARGA COMPLETADA EXITOSAMENTE")
    print("="*100)

except Exception as e:
    cur.execute("ROLLBACK")
    print(f"\n❌ ERROR: {e}")
    import traceback
    traceback.print_exc()

cur.close()
conn.close()
