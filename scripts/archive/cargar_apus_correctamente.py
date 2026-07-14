#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import psycopg2
import sys
sys.stdout.reconfigure(encoding='utf-8')

conn = psycopg2.connect(
    host='localhost',
    database='7_insumos_rado',
    user='postgres',
    password='Jo.9839514500',
    port=5432
)
cur = conn.cursor()

print("="*100)
print("CARGAR APUS_EXTRAIDOS_V2.xlsx CORRECTAMENTE")
print("="*100)

try:
    cur.execute("BEGIN")

    # Limpiar tabla apus_detallado
    print("\n[1/2] Limpiando tabla apus_detallado...")
    cur.execute("DELETE FROM apus_detallado")
    print("  ✓ Limpió")

    # Cargar APUS_Extraidos_v2.xlsx
    print("\n[2/2] Cargando APUS_Extraidos_v2.xlsx...")
    wb = openpyxl.load_workbook('APUS_Extraidos_v2.xlsx')
    ws = wb.active

    cargados = 0
    for row_idx in range(2, ws.max_row + 1):
        try:
            partida_codigo = ws.cell(row_idx, 1).value  # Partida_Codigo
            partida_descripcion = ws.cell(row_idx, 2).value  # Partida_Descripcion
            partida_rendimiento = ws.cell(row_idx, 3).value  # Partida_Rendimiento
            partida_unidad = ws.cell(row_idx, 4).value  # Partida_Unidad
            partida_costo_unitario = ws.cell(row_idx, 5).value  # Partida_Costo_Unitario
            tipo_insumo = ws.cell(row_idx, 6).value  # Tipo_Insumo
            insumo_codigo = ws.cell(row_idx, 7).value  # Insumo_Codigo
            insumo_descripcion = ws.cell(row_idx, 8).value  # Insumo_Descripcion
            insumo_unidad = ws.cell(row_idx, 9).value  # Insumo_Unidad
            insumo_recursos = ws.cell(row_idx, 10).value  # Insumo_Recursos
            insumo_cantidad = ws.cell(row_idx, 11).value  # Insumo_Cantidad
            insumo_precio = ws.cell(row_idx, 12).value  # Insumo_Precio
            insumo_parcial = ws.cell(row_idx, 13).value  # Insumo_Parcial

            if not partida_codigo or not insumo_codigo:
                continue

            cur.execute("""
                INSERT INTO apus_detallado (
                    partida_codigo, partida_descripcion, partida_rendimiento, partida_unidad, partida_costo_unitario,
                    tipo_insumo, insumo_codigo, insumo_descripcion, insumo_unidad, insumo_recursos,
                    insumo_cantidad, insumo_precio, insumo_parcial
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                str(partida_codigo).strip() if partida_codigo else None,
                str(partida_descripcion).strip() if partida_descripcion else None,
                str(partida_rendimiento).strip() if partida_rendimiento else None,
                str(partida_unidad).strip() if partida_unidad else None,
                float(partida_costo_unitario) if partida_costo_unitario else 0,
                str(tipo_insumo).strip() if tipo_insumo else None,
                str(insumo_codigo).strip() if insumo_codigo else None,
                str(insumo_descripcion).strip() if insumo_descripcion else None,
                str(insumo_unidad).strip() if insumo_unidad else None,
                float(insumo_recursos) if insumo_recursos else 0,
                float(insumo_cantidad) if insumo_cantidad else 0,
                float(insumo_precio) if insumo_precio else 0,
                float(insumo_parcial) if insumo_parcial else 0
            ))
            cargados += 1

            if cargados % 500 == 0:
                print(f"  ... {cargados} registros")

        except Exception as e:
            pass

    cur.execute("COMMIT")

    print(f"  ✓ Total cargados: {cargados}")

    # Verificación
    print("\n" + "="*100)
    print("VERIFICACIÓN")
    print("="*100)

    cur.execute("SELECT COUNT(*) FROM apus_detallado")
    total_apus = cur.fetchone()[0]
    print(f"✓ Total registros en apus_detallado: {total_apus}")

    cur.execute("SELECT COUNT(DISTINCT partida_codigo) FROM apus_detallado")
    partidas_unicas = cur.fetchone()[0]
    print(f"✓ Partidas únicas: {partidas_unicas}")

    cur.execute("SELECT COUNT(DISTINCT insumo_codigo) FROM apus_detallado")
    insumos_unicos = cur.fetchone()[0]
    print(f"✓ Insumos únicos: {insumos_unicos}")

    # Ejemplos
    print("\n✓ Ejemplos de datos cargados:")
    cur.execute("""
        SELECT DISTINCT partida_codigo, partida_descripcion
        FROM apus_detallado
        LIMIT 3
    """)
    for cod, desc in cur.fetchall():
        print(f"  {cod}: {desc}")

    print("\n" + "="*100)
    print("✓ APUS CARGADOS CORRECTAMENTE")
    print("="*100)

except Exception as e:
    cur.execute("ROLLBACK")
    print(f"\n❌ ERROR: {e}")
    import traceback
    traceback.print_exc()

cur.close()
conn.close()
