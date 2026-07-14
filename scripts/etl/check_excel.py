import openpyxl

wb = openpyxl.load_workbook('NUEVA_DATA.xlsx', data_only=True)
ws = wb.active
print("Headers:")
headers = [cell.value for cell in ws[1]]
print(headers)
print("First row:")
row2 = [cell.value for cell in ws[2]]
print(row2)
